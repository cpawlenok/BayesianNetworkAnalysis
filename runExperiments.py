import openpyxl as op
import generateBaseCase
from tornadoGraph import getModelOutcomes
import variableElimination
import sys

class experiment:

    def __init__(self, fileName, minFile, maxFile, actions):
        self. fileName = fileName
        self.minFile = minFile
        self.maxFile = maxFile
        self.ogWB = op.load_workbook(self.fileName)  # set up openpyxl workbook for each of the 3 files
        self.minWB = op.load_workbook(self.minFile)  # ^^^ see above comment
        self.maxWB = op.load_workbook(self.maxFile)  # ^^^ see above above comment

    # call to generateBaseCase file
    def runBaseAnalysis(self):
        # Generate the Base Cases
        baseCases = generateBaseCase.generateBaseCases
        baseCases.experimentInit(self.fileName, self.minFile, self.maxFile, self.ogWB, self.minWB, self.maxWB)

    # call to tornadoGraph file
    def runTornadoAnalysis(self, title, termNode, actions):
        getModelOutcomes(numScenarios=16,
                         minWB=self.minWB, maxWB=self.maxWB, ogWB=self.ogWB,
                         title=title, termNode=termNode, actions=actions, outcome="Avoiding Famine State")

def main(JSONFile, Base=False, Tornado=False, monteCarlo=False, actions=[]):
    # newFactors.printFactors()
    # main contains all of the file declarations
    # main arguments all set to False by default to prevent the program from running way longer than needed
    # Set up file paths here for base file, and destination files for min and max cases
    fileName1 = "excelFiles/Model_3T_Drought_Data.xlsx"  # file we are pulling data from to generate min and max models
    minFile1 = "excelFiles/minDroughtModels.xlsx"  # where we are placing the min generated models
    maxFile1 = "excelFiles/maxDroughtModels.xlsx"  # where we are placing the max generated models

    fileName2 = "excelFiles/Model_3T_Rainfall_Data.xlsx"
    minFile2 = "excelFiles/minRainfallModels.xlsx"
    maxFile2 = "excelFiles/maxRainfallModels.xlsx"

    fileName3 = "excelFiles/Model_3T_v1Data.xlsx"
    minFile3 = "excelFiles/minModels.xlsx"
    maxFile3 = "excelFiles/maxModels.xlsx"

    # Set up experiment objects for all experiments
    experiment_1 = experiment(fileName1, minFile1, maxFile1, actions)
    experiment_2 = experiment(fileName2, minFile2, maxFile2, actions)
    experiment_3 = experiment(fileName3, minFile3, maxFile3, actions)

    variableElimination.startElim(JSONFile, experiment_2, actions)
    Tornado = True
    # run tornado analysis - variable elimination has to be performed first
    if Tornado == True:
        # termNode sets the output node of the network
        # termNode is the excel column of the terminal Node
        experiment_1.runTornadoAnalysis(title='Full Case', termNode=16, actions=actions)
        # experiment_2.runTornadoAnalysis(title='Rainfall Case', termNode=10)
        # experiment_3.runTornadoAnalysis(title='Full Case', termNode=16)

    # if monteCarlo == True:

    # now we must perform variable elimination
    # use tornadoGraph to graph the obtain and graph the results

if __name__ == '__main__':
    # collect JSON Filepath from the command line
    # JSONFile = str(sys.argv[1])

    # start program
    main(JSONFile="C:/Users/eleme/Documents/School Folders/Research Assitantship Files/Project Folder 2.0/Pythia/Pythia 1.8_v2/Pythia 1.8/JSONNetworkData.json",
         actions=["Food Imports", "Direct Aid", "Food Aid Provision", "Conflict Resolution"])

