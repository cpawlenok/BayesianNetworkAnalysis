import openpyxl as op
from matplotlib import pyplot as plt


# collects all of the data from models after variable elimination
# then produces tornado graph using matplotlib
# arguments:
#       totalScenarios - total number of scenarios = (# of decision nodes)^2
#       scenarioNum - number of scenario while iterating through scenarios
#       minWS - the excel worksheet object created by openpyxl for min case
#       maxWS - the excel worksheet object created by openpyxl for max case
class Scenario:
    def __init__(self, totalScenarios, scenarioNum, minWS, maxWS):
        self.scenarioNum = scenarioNum
        self.minWS = minWS
        self.maxWS = maxWS
        self.totalScenarios = totalScenarios

    def getMinCases(self, termNode):
        minXvals = []
        minYvals = []
        modelNumber = 0
        for col in self.minWS.iter_cols(min_col=3, values_only=True):

            if modelNumber == 0:
                minXvals.append(self.minWS.cell(column=termNode, row=self.scenarioNum+2).value)
                minYvals.append(col[0])
            else:
                minXvals.append(self.minWS.cell(column=termNode, row=2+((modelNumber*self.totalScenarios)+self.scenarioNum)).value)
                minYvals.append(col[0])
            modelNumber += 1
        return [minXvals, minYvals]

    def getMaxCases(self, termNode):
        maxXvals = []
        maxYvals = []
        modelNumber = 0
        for col in self.maxWS.iter_cols(min_col=3, values_only=True):

            if modelNumber == 0:
                maxXvals.append(self.maxWS.cell(column=termNode, row=self.scenarioNum + 2).value)
                maxYvals.append(col[0])
            else:
                maxXvals.append(self.maxWS.cell(column=termNode, row=2+((modelNumber*self.totalScenarios)+self.scenarioNum)).value)
                maxYvals.append(col[0])
            modelNumber += 1
        return [maxXvals, maxYvals]


def getBaseCases(baseWS, scenarioNum, termNode):
    baseXvals = []
    baseYvals = []
    for col in baseWS.iter_cols(min_col=3, values_only=True):
        baseXvals.append(baseWS.cell(column=termNode, row=scenarioNum+2).value)
        baseYvals.append(col[0])
    return [baseXvals, baseYvals]

def sortModel(min, base, max):
    cases = []
    for x in range(len(min[0])):
        cases.append(abs(min[0][x]-max[0][x]))
    length = len(min[0])
    for i in range(length-1):
        for j in range(0, length-i-1):

            if cases[j] > cases[j+1]:

                min[0][j], min[0][j+1] = min[0][j+1], min[0][j]
                min[1][j], min[1][j + 1] = min[1][j+1], min[1][j]
                base[0][j], base[0][j+1] = base[0][j+1], base[0][j]
                base[1][j], base[1][j + 1] = base[1][j+1], base[1][j]
                max[0][j], max[0][j+1] = max[0][j+1], max[0][j]
                max[1][j], max[1][j + 1] = max[1][j+1], max[1][j]

                cases[j], cases[j+1] = cases[j+1], cases[j]

    return min, base, max

def plotTornado(title, case):
    minCasesUnsorted = case[0]
    baseCasesUnsorted = case[1]
    maxCasesUnsorted = case[2]
    y = range(len(minCasesUnsorted[0]))
    y = list(y)
    minCases, baseCases, maxCases = sortModel(minCasesUnsorted, baseCasesUnsorted, maxCasesUnsorted)

    for index in range(len(y)):
        minX = minCases[0][index]
        baseX = baseCases[0][index]
        maxX = maxCases[0][index]

        if minX>=baseX and maxX>=baseX:
            minLength = abs(minX - baseX)
            maxLength = abs(maxX - baseX)
            plt.broken_barh(
                [(baseX, minLength), (baseX, maxLength)],
                (y[index] - 0.3, 0.8),
                facecolor='blue',
                edgecolor='black',
                linewidth=1
            )

        if minX<=baseX and maxX>=baseX:
            minLength = abs(minX - baseX)
            maxLength = abs(maxX - baseX)
            plt.broken_barh(
                [(minX, minLength), (baseX, maxLength)],
                (y[index] - 0.3, 0.8),
                facecolor='blue',
                edgecolor='black',
                linewidth=1
            )

        if minX >= baseX and maxX <= baseX:
            minLength = abs(minX - baseX)
            maxLength = abs(maxX - baseX)
            plt.broken_barh(
                [(maxX, maxLength), (baseX, minLength)],
                (y[index] - 0.3, 0.8),
                facecolor='blue',
                edgecolor='black',
                linewidth=1
            )

        if minX <= baseX and maxX <= baseX:
            minLength = abs(minX - baseX)
            maxLength = abs(maxX - baseX)
            plt.broken_barh(
                [(minX, minLength), (maxX, maxLength)],
                (y[index] - 0.3, 0.8),
                facecolor='blue',
                edgecolor='black',
                linewidth=1
            )


        plt.axvline(baseCases[0][1], color='black')

        axes = plt.gca()
        axes.spines['left'].set_visible(False)
        axes.spines['right'].set_visible(False)
        axes.spines['bottom'].set_visible(False)
        axes.xaxis.set_ticks_position('top')

        plt.title(title)
        plt.xlabel("Outcome Probability")
        plt.ylabel("Node")
        plt.yticks(y, baseCases[1])
        plt.xlim(baseCases[0][0] - 0.25, baseCases[0][0] + 0.25)
        plt.ylim(-1, len(y))
    plt.show()

# Main function of the file, this function calls all of the other functions to generate tornado diagrams
def getModelOutcomes(numScenarios, minWB, maxWB, ogWB, title, termNode):
    baseWS = ogWB["Sheet3"]
    minWS = minWB["Sheet3"]
    maxWS = maxWB["Sheet3"]

    indexedMinCases = []
    indexedMaxCases = []
    indexedBaseCases = []

    for i in range(numScenarios):
        scenario = Scenario(numScenarios, i, minWS, maxWS)
        minCases = scenario.getMinCases(termNode)
        maxCases = scenario.getMaxCases(termNode)

        indexedMinCases.append(minCases)
        indexedBaseCases.append(getBaseCases(baseWS, i, termNode))
        indexedMaxCases.append(maxCases)

    for i in range(numScenarios):
        case = (indexedMinCases[i], indexedBaseCases[i], indexedMaxCases[i])
        plotTitle = title + " Scenario Number: " + str(i)
        plotTornado(plotTitle, case)
