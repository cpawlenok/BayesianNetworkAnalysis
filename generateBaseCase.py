import openpyxl as op

class Model:
    def __init__(self, rows, columns):
        self.rows = rows
        self.columns = columns
        self.data = []

    # adds column to the list of data
    def addColumn(self, data):
        self.data.append(data)

    # only used for testing purposes to confirm results
    def printData(self):
        for column in self.data:
            print(column)

    # accesses column at the appropriate index
    # index is
    def getColumn(self, index):
        return self.data[index][:]

    def getRow(self, index):
        rowData = []
        for i in range(self.columns):
            rowData.append(self.data[index][i])
        return rowData

    def setMin(self, cell):
        self.data[cell[0]][cell[1]] = 0.0

    # Receives Cell (col, row)
    def setMax(self, cell):
        self.data[cell[0]][cell[1]] = 1


def generateModels(wb):
    # open excel workbook and collect correct sheet information
    ws = wb["Sheet2"]
    numRows = len(ws["A"])
    numColumns = len(ws[1])
    models = []

    # generate models
    for i in range(numRows-1):                      # we need 37 distinct models
        currModel = Model(numRows, numColumns)      # create new model
        # get columns from worksheet and add them to the Model object
        for col in ws.iter_cols(min_row=2, values_only=True):
            column = list(col)
            currModel.addColumn(column)
        models.append(currModel)

    return models

# iterates through models to set correct cells to 0.0
# uses the model class function setMin to do so
def minModels(models):
    rowNum = 0
    for model in models:
        cell = (2, rowNum)
        model.setMin(cell)
        cell2 = (1, rowNum)
        model.setMin(cell2)
        rowNum += 1

    return models

# iterates through models to set correct cells to 1.0
# uses the model class function setMax to do so
def maxModels(models):
    rowNum = 0
    for model in models:
        cell = (2, rowNum)
        model.setMax(cell)
        cell2 = (1, rowNum)
        model.setMax(cell2)
        rowNum += 1

    return models

def createModelWS(min, models, wb, ogWB, fileName):
    ws1 = wb["Sheet1"]
    ws2 = wb["Sheet2"]


    sheetOneWS = ogWB["Sheet1"]
    sheetOneRows = len(sheetOneWS["A"])
    sheetOneCols = len(sheetOneWS[1])
    sheetOneModel = Model(sheetOneRows, sheetOneCols)

    for col in sheetOneWS.iter_cols(min_row=2, values_only=True):
        column = list(col)
        sheetOneModel.addColumn(column)

    numRows = len(models[0].getColumn(0))
    numCols = len(models[0].getRow(0))
    modelNum = 0

    for model in models:
        for i in range(numRows):
            for j in range(numCols):
                # print(model.data)
                if modelNum==0:
                    ws2.cell(column=j+1, row=2+i).value = model.data[j][i]
                else:
                    ws2.cell(column=j+1, row=2+i+(modelNum*(numRows+1))).value = model.data[j][i]
            if min:
                wb.save(fileName)
            else:
                wb.save(fileName)
        modelNum += 1

    # since Pythia requires both sheet one and sheet two to have the same number of models
    # this creates sheet one of the project which specifies g and h values of each link in the network
    # note: each model should be exactly the same

    # todo: this is irrelevant after variable elimination is implemented in pythonScripts
    oneNumRows = len(sheetOneModel.getColumn(0))
    oneNumCols = len(sheetOneModel.getRow(0))
    oneModelNum = 0
    for index in range(numRows):
        for i in range(oneNumRows):
            for j in range(oneNumCols):
                if oneModelNum == 0:
                    ws1.cell(column=j+1, row=2+i).value = sheetOneModel.data[j][i]
                else:
                    ws1.cell(column=j+1, row=2+i+(oneModelNum*(oneNumRows+1))).value = sheetOneModel.data[j][i]
            if min:
                wb.save(fileName)
            else:
                wb.save(fileName)
        oneModelNum += 1


def clearSheet(WB, fileName):
    ws1 = WB["Sheet1"]
    ws2 = WB["Sheet2"]
    ws3 = WB["Sheet3"]

    ws1.delete_rows(1, ws1.max_row+1)
    ws2.delete_rows(1, ws2.max_row+1)
    ws3.delete_rows(1, ws3.max_row+1)

    if min:
        WB.save(fileName)
    else:
        WB.save(fileName)

class generateBaseCases:

    def experimentInit(self, minFile, maxFile, ogWB, minWB, maxWB):
        # spring cleaning
        clearSheet(minWB, minFile)
        clearSheet(maxWB, maxFile)

        # generate min models
        # store min models in separate variable from max models
        models4min = generateModels(ogWB)
        minModel = minModels(models4min)

        # generate max models
        # store max models in separate variable from min models
        models4max = generateModels(ogWB)
        maxModel = maxModels(models4max)

        # output the models to excel
        # each output goes to corresponding excel file
        min = True
        createModelWS(min, minModel, minWB, ogWB, minFile)
        min = False
        createModelWS(min, maxModel, maxWB, ogWB, maxFile)